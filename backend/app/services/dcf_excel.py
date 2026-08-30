"""DCF xlsx writer — render a DcfPayload to an Excel workbook.

Multi-tab structure:
- Cover            — case metadata + EU funding statement
- Actors           — header + whatever the Network Builder stored, then
                     blank rows for offline collection
- Flow Matrix      — same
- Logistics        — same (or placeholder note if section
                     deactivated, e.g. q7=A)
- Infrastructure   — header + 10 empty rows
- Methodological Choices — auto-populated checklist of activated
                     procedural_mandates, grouped by category
- Network Diagram  — placeholder note (the interactive viz is in-app)

Design principles:
- SYMBA palette: blue header (#1F4E79) + white text, italic small gray
  for field-id technical reference, EU footer on every sheet (inline + page
  footer for print).
- No macros, no VBA — Excel/LibreOffice/Google Sheets compatible.
- Returns bytes; caller decides how to serve (HTTP / file / etc).
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.domain.dcf_data import DcfData, DcfRow
from app.engine.dcf_compose import DcfPayload, DcfSection
from app.engine.dcf_rows import cell_values, edge_list, rows_for

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EU_FOOTER = (
    "This Project has received funding from the European Union's Horizon "
    "Research and Innovation Programme under Grant Agreement N. 101135562 — "
    "www.symbaproject.eu"
)

# SYMBA palette: white + blue + green geometric (per project visual identity).
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
_SUBTITLE_FONT = Font(italic=True, size=10, color="404040")
_FIELDID_FONT = Font(size=8, italic=True, color="666666")
_LABEL_FONT = Font(bold=True)
_FOOTER_FONT = Font(size=8, italic=True, color="888888")

_DATA_SECTION_ORDER = [
    "actors", "flow_matrix", "logistics", "costs", "infrastructure",
]
_SHEET_NAMES = {
    "actors": "Actors",
    "flow_matrix": "Flow Matrix",
    "logistics": "Logistics",
    "costs": "Costs & Revenues",
    "infrastructure": "Infrastructure",
}
_EMPTY_DATA_ROWS = 10
# Even a fully-filled section keeps a few blank rows: collection continues
# offline after the export.
_MIN_BLANK_ROWS = 3


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def render_xlsx(
    payload: DcfPayload,
    data: DcfData | None = None,
    case_title: str | None = None,
) -> bytes:
    """Render the DcfPayload to an xlsx workbook and return raw bytes.

    `data` is the content stored by the Network Builder, when the case
    has any: its rows are written under the headers so the analyst gets
    a partially-filled workbook to continue from, instead of the empty
    grid the export produced before.
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    _write_cover_tab(wb, payload, case_title)
    _write_instructions_tab(wb, payload, case_title)

    sections = {s.id: s for s in payload.sections}

    for sec_id in _DATA_SECTION_ORDER:
        sec = sections.get(sec_id)
        if sec is not None:
            _write_data_section_tab(wb, sec, rows_for(data, sec_id))

    mc = sections.get("methodological_choices")
    if mc is not None:
        _write_mandates_tab(wb, mc, payload.obligations)

    nw = sections.get("network_diagram")
    if nw is not None:
        _write_network_tab(wb, nw, data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tab writers
# ---------------------------------------------------------------------------


def _write_cover_tab(
    wb: Workbook, payload: DcfPayload, case_title: str | None = None
) -> None:
    ws = wb.create_sheet("Cover")
    # The recipient knows the case by its name, not by a uuid.
    ws["A1"] = (
        f"SYMBA T4.6 — Data Collection File — {case_title}"
        if case_title
        else "SYMBA T4.6 — Data Collection File"
    )
    ws["A1"].font = _TITLE_FONT

    rows = [
        ("Case ID", payload.case_id),
        ("Pathway", payload.pathway_id or "—"),
        ("ILCD Situation", payload.ilcd_situation or "—"),
        ("LCC Type", payload.lcc_type or "—"),
        ("S-LCA Activation", payload.slca_activation_state or "—"),
        ("IS-01 Extended", "yes" if payload.is_01_extended else "no"),
        ("Schema version", payload.schema_version),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=i, column=2, value=value)

    footer_row = 3 + len(rows) + 2
    ws.cell(row=footer_row, column=1, value=EU_FOOTER).font = _FOOTER_FONT
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=4)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50
    _set_print_footer(ws)


def _write_data_section_tab(
    wb: Workbook, section: DcfSection, rows: list[DcfRow] | None = None
) -> None:
    ws = wb.create_sheet(_SHEET_NAMES[section.id])

    if not section.active:
        ws["A1"] = section.title_en
        ws["A1"].font = _TITLE_FONT
        ws["A3"] = (
            f"This section is NOT activated for the current case "
            f"(applies_when: {section.applies_when})."
        )
        ws["A3"].alignment = Alignment(wrap_text=True)
        ws.cell(row=5, column=1, value=EU_FOOTER).font = _FOOTER_FONT
        ws.column_dimensions["A"].width = 80
        _set_print_footer(ws)
        return

    ws["A1"] = section.title_en
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = section.description_en
    ws["A2"].font = _SUBTITLE_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")

    # Row 4: header labels. The technical field id and the description
    # ride along as a cell comment instead of a second visible row: the
    # sheet has to read as a form to whoever fills it in, while the id
    # stays available to whoever maps the file back to the schema.
    header_row = 4
    for col_idx, field in enumerate(section.fields, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=_header_label(field))
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.comment = Comment(_field_note(field), "SYMBA T4.6")

    fieldid_row = header_row

    # Stored rows first, then blank ones so collection can continue.
    stored = rows or []
    first_data_row = fieldid_row + 1
    for offset, row in enumerate(stored):
        for col_idx, value in enumerate(cell_values(section, row), start=1):
            ws.cell(row=first_data_row + offset, column=col_idx, value=value)

    blank_start = first_data_row + len(stored)
    blank_count = max(_EMPTY_DATA_ROWS - len(stored), _MIN_BLANK_ROWS)
    for r in range(blank_start, blank_start + blank_count):
        for col_idx in range(1, len(section.fields) + 1):
            ws.cell(row=r, column=col_idx, value=None)
    total_data_rows = len(stored) + blank_count

    _add_enum_dropdowns(ws, section, first_data_row, blank_start + blank_count)

    # Column widths
    for col_idx, field in enumerate(section.fields, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(15, min(36, len(field.label_en) + 4))

    # Freeze top 5 rows (title + description + 2 header rows + small buffer)
    ws.freeze_panes = ws.cell(row=fieldid_row + 1, column=1).coordinate

    # Footer
    footer_row = fieldid_row + 1 + total_data_rows + 2
    ws.cell(row=footer_row, column=1, value=EU_FOOTER).font = _FOOTER_FONT

    _set_print_footer(ws)


_INSTRUCTIONS = [
    ("What this file is",
     "A data collection worksheet generated for one industrial-symbiosis "
     "case. The columns were selected by the assessment's methodological "
     "pathway: you are only asked for what this case needs."),
    ("How to fill it in",
     "One row per item. Columns marked * are required. Hover a column "
     "header for the definition, the allowed values and the reason the "
     "field is asked. Cells with a dropdown only accept the listed values "
     "— do not overwrite them with free text."),
    ("If you do not know a value",
     "Leave it empty rather than guessing, and say so to the analyst. An "
     "empty cell is a known gap; a guessed number is an unknown error."),
    ("Units and currency",
     "Always state the unit next to the quantity where a unit column "
     "exists, and use the currency and reference year the Costs sheet "
     "asks for. Figures without a unit cannot be aggregated."),
    ("Confidentiality",
     "This workbook may carry data from every partner in the network. "
     "Before circulating it outside the assessment team, check the "
     "confidentiality arrangement between the partners — the "
     "methodology requires it to be settled before data collection "
     "starts."),
    ("Who to return it to",
     "The analyst who sent you this file. Agree a date with them: an "
     "assessment stalls on the last missing sheet."),
]


def _write_instructions_tab(
    wb: Workbook, payload: DcfPayload, case_title: str | None = None
) -> None:
    """A sheet for the person who has to fill this in.

    The workbook used to arrive with no instructions, no deadline and no
    contact — a recipient outside the assessment team had no way of
    knowing what was expected of them.
    """
    ws = wb.create_sheet("Instructions")
    ws["A1"] = "How to fill in this Data Collection File"
    ws["A1"].font = _TITLE_FONT
    if case_title:
        ws["A2"] = f"Case: {case_title}"
        ws["A2"].font = _SUBTITLE_FONT

    row = 4
    for heading, body in _INSTRUCTIONS:
        c = ws.cell(row=row, column=1, value=heading)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        ws.cell(row=row, column=2, value=body).alignment = Alignment(wrap_text=True)
        row += 2

    ws.cell(row=row, column=1, value="Return by").font = _HEADER_FONT
    ws.cell(row=row, column=2, value="(agree a date with the analyst)")
    row += 1
    ws.cell(row=row, column=1, value="Contact").font = _HEADER_FONT
    ws.cell(row=row, column=2, value="(analyst name and e-mail)")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    ws.cell(row=row + 3, column=1, value=EU_FOOTER).font = _FOOTER_FONT
    _set_print_footer(ws)


def _header_label(field) -> str:
    """Column header as a person reads it: the label, and a star when the
    field has to be filled in."""
    return f"{field.label_en} *" if field.required else field.label_en


def _field_note(field) -> str:
    parts = [f"Field id: {field.id}", f"Type: {field.type}"]
    if field.description_en:
        parts.append(str(field.description_en))
    if field.enum_values:
        parts.append("Allowed: " + ", ".join(field.enum_values))
    if field.activation_predicate and field.activation_predicate != "always":
        parts.append(f"Asked because: {field.activation_predicate}")
    return "\n".join(parts)


def _add_enum_dropdowns(ws, section: DcfSection, first_row: int, last_row: int) -> None:
    """Turn every inline enum into a dropdown.

    Typing an allowed value from memory is how a collection file comes
    back with 'medium', 'Med' and 'M' in the same column.
    """
    for col_idx, field in enumerate(section.fields, start=1):
        values = field.enum_values or []
        if not values:
            continue
        joined = ",".join(values)
        # Excel caps an inline list at 255 characters; a longer vocabulary
        # would silently break the sheet, so it stays free text.
        if len(joined) > 250:
            continue
        dv = DataValidation(
            type="list", formula1=f'"{joined}"', allow_blank=True, showDropDown=False
        )
        dv.error = "Pick one of the listed values."
        dv.errorTitle = "Value not allowed"
        ws.add_data_validation(dv)
        letter = get_column_letter(col_idx)
        dv.add(f"{letter}{first_row}:{letter}{last_row}")


def _write_mandates_tab(
    wb: Workbook,
    section: DcfSection,
    obligations: list,
) -> None:
    """One tab for everything the case has to document.

    Procedural mandates and triggered cross-method rules used to live in
    two different places in two different formats; they are the same job
    for whoever fills the file in, so they share a tab, with `Origin`
    keeping the traceability a reviewer needs.
    """
    ws = wb.create_sheet("Methodological Choices")
    ws["A1"] = section.title_en
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = (
        "FOR THE ASSESSMENT TEAM — not part of what a data provider fills "
        "in. " + section.description_en
    )
    ws["A2"].font = _SUBTITLE_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:I2")

    headers = [
        "Category", "Origin", "ID", "Method", "Statement", "Applies because",
        "Deliverable target", "Assignee", "Status",
    ]
    header_row = 4
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = header_row + 1
    for item in obligations:
        ws.cell(row=row, column=1, value=item.category)
        ws.cell(row=row, column=2, value=item.origin)
        ws.cell(row=row, column=3, value=item.id)
        ws.cell(row=row, column=4, value=item.method)
        statement = item.statement if not item.title else f"{item.title} — {item.statement}"
        ws.cell(row=row, column=5, value=statement).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=6, value=item.trigger or "")
        ws.cell(row=row, column=7, value="")
        ws.cell(row=row, column=8, value="")
        ws.cell(row=row, column=9, value="pending")
        row += 1

    widths = {"A": 24, "B": 10, "C": 14, "D": 8, "E": 60, "F": 26,
              "G": 28, "H": 18, "I": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A5"
    ws.cell(row=row + 2, column=1, value=EU_FOOTER).font = _FOOTER_FONT
    _set_print_footer(ws)


def _write_network_tab(
    wb: Workbook, section: DcfSection, data: DcfData | None
) -> None:
    """Edge list of the drawn network, or the placeholder note when the
    analyst has not drawn one yet."""
    ws = wb.create_sheet("Network Diagram")
    ws["A1"] = section.title_en
    ws["A1"].font = _TITLE_FONT

    edges = edge_list(data)
    if not edges:
        ws["A3"] = (
            "No network drawn yet. Use the Network Builder in the SYMBA T4.6 web "
            "app ('Data Collection' page) to place actors and connect them — the "
            "edges then appear here, and the Actors / Flow Matrix tabs come "
            "pre-filled."
        )
        ws["A3"].alignment = Alignment(wrap_text=True)
        ws.column_dimensions["A"].width = 90
        ws.cell(row=5, column=1, value=EU_FOOTER).font = _FOOTER_FONT
        _set_print_footer(ws)
        return

    ws["A2"] = (
        "Edge list of the network drawn in the app. The interactive diagram "
        "itself lives on the 'Data Collection' page."
    )
    ws["A2"].font = _SUBTITLE_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)

    headers = ["Flow", "Origin actor", "Destination actor", "Type", "Quantity", "Unit"]
    header_row = 4
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for offset, edge in enumerate(edges, start=1):
        ws.cell(row=header_row + offset, column=1, value=edge["name"])
        ws.cell(row=header_row + offset, column=2, value=edge["origin"])
        ws.cell(row=header_row + offset, column=3, value=edge["destination"])
        ws.cell(row=header_row + offset, column=4, value=edge["type"])
        ws.cell(row=header_row + offset, column=5, value=edge["quantity"])
        ws.cell(row=header_row + offset, column=6, value=edge["unit"])

    for col_idx, label in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(label) + 6)

    ws.cell(
        row=header_row + len(edges) + 2, column=1, value=EU_FOOTER
    ).font = _FOOTER_FONT
    _set_print_footer(ws)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _set_print_footer(ws) -> None:
    """Best-effort print footer (visible when sheet is printed)."""
    try:
        ws.oddFooter.center.text = EU_FOOTER
        ws.oddFooter.center.size = 8
    except Exception:
        # openpyxl HeaderFooter API can be flaky across versions; the inline
        # footer cell is the authoritative one. Swallow.
        pass
