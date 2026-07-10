"""Tests for app.services.dcf_excel.

Verifies the rendered workbook is structurally correct (right tabs,
header cells, expected technical-id row, EU footer present).
"""
from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.domain.enums import Q1, Q2, Q4, Q5, Q7, Q6a, Q6b
from app.domain.models import Q3, Case, Flow
from app.engine.dcf_compose import compose_dcf
from app.engine.pipeline import run as pipeline_run
from app.services.dcf_excel import EU_FOOTER, render_xlsx


@pytest.fixture(scope="module")
def dcf_schema() -> dict:
    p = (
        Path(__file__).resolve().parents[1] / "app" / "schemas" / "dcf_schema.json"
    )
    with open(p) as f:
        return json.load(f)


@pytest.fixture(scope="module")
def mandates_census() -> dict:
    p = (
        Path(__file__).resolve().parents[1]
        / "coordination" / "dcf_mandates_census.json"
    )
    with open(p) as f:
        return json.load(f)


def _flows(*qs):
    return [Flow(id=f"f{i+1}", name=f"flow{i+1}", q5=q) for i, q in enumerate(qs)]


@pytest.fixture()
def payload_wiktor(schemas, dcf_schema, mandates_census):
    case = Case(
        q1=Q1.B, q2=Q2.D, q3=Q3(env=True, eco=True), q4={Q4.E},
        q6a=Q6a.WASTEWATER_SLUDGE_BIOFACTORIES, q6b=Q6b.TRL7_8, q7=Q7.B,
        flows=_flows(Q5.a, Q5.c, Q5.c),
    )
    pipeline_run(case, schemas)
    return compose_dcf(case, dcf_schema, mandates_census)


@pytest.fixture()
def payload_arce(schemas, dcf_schema, mandates_census):
    """q7=A → logistics tab will be a placeholder."""
    case = Case(
        q1=Q1.B, q2=Q2.A, q3=Q3(env=True), q4={Q4.E},
        q6a=Q6a.PLASTICS_PACKAGING, q6b=Q6b.TRL9, q7=Q7.A,
        flows=_flows(Q5.b, Q5.c),
    )
    pipeline_run(case, schemas)
    return compose_dcf(case, dcf_schema, mandates_census)


# ---------------------------------------------------------------------------
# Structure
# ---------------------------------------------------------------------------


def test_render_returns_bytes(payload_wiktor):
    blob = render_xlsx(payload_wiktor)
    assert isinstance(blob, bytes)
    assert len(blob) > 1000  # workbook isn't empty


def test_workbook_has_expected_tabs(payload_wiktor):
    blob = render_xlsx(payload_wiktor)
    wb = load_workbook(BytesIO(blob))
    expected = ["Cover", "Actors", "Flow Matrix", "Logistics",
                "Infrastructure", "Methodological Choices", "Network Diagram"]
    assert wb.sheetnames == expected


def test_cover_tab_carries_case_metadata(payload_wiktor):
    blob = render_xlsx(payload_wiktor)
    wb = load_workbook(BytesIO(blob))
    cover = wb["Cover"]
    assert cover["A1"].value == "SYMBA T4.6 — Data Collection File"

    # Read the label/value pairs from rows 3+
    labels = {cover.cell(row=r, column=1).value for r in range(3, 11)}
    assert "Case ID" in labels
    assert "Pathway" in labels
    assert "ILCD Situation" in labels


def test_pathway_value_in_cover(payload_wiktor):
    blob = render_xlsx(payload_wiktor)
    wb = load_workbook(BytesIO(blob))
    cover = wb["Cover"]
    # find row with "Pathway"
    for r in range(3, 11):
        if cover.cell(row=r, column=1).value == "Pathway":
            assert cover.cell(row=r, column=2).value == "IS-01"
            break
    else:
        pytest.fail("Pathway row not found")


# ---------------------------------------------------------------------------
# Data section tabs
# ---------------------------------------------------------------------------


def test_actors_tab_has_header_row(payload_wiktor):
    blob = render_xlsx(payload_wiktor)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Actors"]
    # Row 4 should be the header labels
    header_values = [ws.cell(row=4, column=c).value for c in range(1, 10)]
    # at least "Actor ID" should appear
    assert "Actor ID" in header_values


def test_actors_tab_has_field_id_subheader(payload_wiktor):
    blob = render_xlsx(payload_wiktor)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Actors"]
    # Row 5 should be the technical field IDs
    ids = [ws.cell(row=5, column=c).value for c in range(1, 10)]
    assert "actor.id" in ids
    assert "actor.role" in ids


def test_flow_matrix_tab_has_marginal_market_when_q1c(
    schemas, dcf_schema, mandates_census
):
    """q1=C → flow.marginal_market_ref should be in the header."""
    case = Case(
        q1=Q1.C, q2=Q2.C, q3=Q3(env=True, eco=True, soc=True), q4={Q4.E},
        q6a=Q6a.BIOBASED_POLYMERS, q6b=Q6b.TRL9, q7=None,
        flows=_flows(Q5.b, Q5.c),
    )
    pipeline_run(case, schemas)
    payload = compose_dcf(case, dcf_schema, mandates_census)
    blob = render_xlsx(payload)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Flow Matrix"]
    # field IDs row 5
    ids = [ws.cell(row=5, column=c).value for c in range(1, 30) if ws.cell(row=5, column=c).value]
    assert "flow.marginal_market_ref" in ids


def test_logistics_inactive_tab_shows_placeholder(payload_arce):
    blob = render_xlsx(payload_arce)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Logistics"]
    # Should NOT have a normal header row at row 4
    a3 = ws["A3"].value
    assert a3 is not None and "NOT activated" in a3


# ---------------------------------------------------------------------------
# Methodological choices tab
# ---------------------------------------------------------------------------


def test_mandates_tab_has_mandate_rows(payload_wiktor):
    blob = render_xlsx(payload_wiktor)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Methodological Choices"]
    # Header row at 4
    headers = [ws.cell(row=4, column=c).value for c in range(1, 8)]
    assert headers == [
        "Category", "Node ID", "Method", "Statement",
        "Deliverable target", "Assignee", "Status",
    ]
    # At least 50 mandate rows
    row_count = 0
    for r in range(5, 200):
        if ws.cell(row=r, column=2).value:  # node_id column
            row_count += 1
    assert row_count >= 50


# ---------------------------------------------------------------------------
# EU footer
# ---------------------------------------------------------------------------


def test_eu_footer_present_in_cover(payload_wiktor):
    blob = render_xlsx(payload_wiktor)
    wb = load_workbook(BytesIO(blob))
    cover = wb["Cover"]
    # Footer somewhere in column A
    has_footer = False
    for r in range(1, 30):
        v = cover.cell(row=r, column=1).value
        if v and "Grant Agreement N. 101135562" in str(v):
            has_footer = True
            break
    assert has_footer


def test_eu_footer_text_is_canonical(payload_wiktor):
    blob = render_xlsx(payload_wiktor)
    wb = load_workbook(BytesIO(blob))
    # canonical reference is in our module
    assert "101135562" in EU_FOOTER
    assert "www.symbaproject.eu" in EU_FOOTER
    assert "Horizon" in EU_FOOTER
    # check each tab has the footer in column A somewhere
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        found = False
        # max_row is variable; cap at 200
        for r in range(1, min(ws.max_row + 1, 200)):
            v = ws.cell(row=r, column=1).value
            if v and "101135562" in str(v):
                found = True
                break
        assert found, f"EU footer missing in tab {sheet_name!r}"
